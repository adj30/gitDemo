import openpyxl

book = openpyxl.load_workbook("D:\\pythonSeleniumData.xlsx")
sheet1 = book.active
cell = sheet1.cell(row=2, column=1)
print(cell.value)
sheet1.cell(row=3, column=2).value = "Aditya"
print(sheet1.cell(row=3, column=2).value)

print(sheet1.max_row)
print(sheet1.max_column)
print(sheet1["C7"].value)

Dict = {}

for i in range(1, sheet1.max_row + 1):  # to get rows
    if sheet1.cell(row=i, column=1).value == "TestCase5":

        for j in range(2, sheet1.max_column + 1):  # to get columns
            print(sheet1.cell(row=i, column=j).value, end=" ")
            Dict[sheet1.cell(row=1, column=j).value] = sheet1.cell(row=i, column=j).value

print(Dict)
