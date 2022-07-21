import openpyxl


class HomepageData:
    test_Homepage_data = [{"firstName": "Sachin", "email": "s@gmail.com", "password": "admin1", "gender": "Male"},
                          {"firstName": "Rahul", "email": "r@gmail.com", "password": "admin2", "gender": "Male"},
                          {"firstName": "Mithali", "email": "m@gmail.com", "password": "admin3", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("D:\\pythonSeleniumData_2.xlsx")
        sheet1 = book.active
        for i in range(1, sheet1.max_row + 1):  # to get rows
            if sheet1.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet1.max_column + 1):  # to get columns
                    print(sheet1.cell(row=i, column=j).value, end=" ")
                    Dict[sheet1.cell(row=1, column=j).value] = sheet1.cell(row=i, column=j).value
        return [Dict]
